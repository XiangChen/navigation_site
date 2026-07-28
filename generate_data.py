import openpyxl
import json
import os

# Read all product data
wb = openpyxl.load_workbook('aixzd_ai产品(1).xlsx')

categories = {
    'aixzd_AI写作工具': 'writing',
    'aixzd_AI创意工具': 'creative',
    'aixzd_AI办公工具': 'office',
    'aixzd_AI图形处理': 'image',
    'aixzd_AI学习平台': 'learning',
    'aixzd_AI对话聊天': 'chat',
    'aixzd_AI工具箱': 'toolbox',
    'aixzd_AI智能助手': 'assistant',
    'aixzd_AI绘画工具': 'drawing',
    'aixzd_AI编程工具': 'coding',
    'aixzd_AI翻译工具': 'translation',
    'aixzd_AI营销工具': 'marketing',
    'aixzd_AI视频工具': 'video',
    'aixzd_AI设计工具': 'design',
    'aixzd_AI音频工具': 'audio'
}

category_names = {
    'writing': 'AI写作工具',
    'creative': 'AI创意工具',
    'office': 'AI办公工具',
    'image': 'AI图形处理',
    'learning': 'AI学习平台',
    'chat': 'AI对话聊天',
    'toolbox': 'AI工具箱',
    'assistant': 'AI智能助手',
    'drawing': 'AI绘画工具',
    'coding': 'AI编程工具',
    'translation': 'AI翻译工具',
    'marketing': 'AI营销工具',
    'video': 'AI视频工具',
    'design': 'AI设计工具',
    'audio': 'AI音频工具'
}

# Category icons
category_icons = {
    'writing': 'edit_square',
    'creative': 'lightbulb',
    'office': 'work',
    'image': 'image_search',
    'learning': 'school',
    'chat': 'chat',
    'toolbox': 'handyman',
    'assistant': 'smart_toy',
    'drawing': 'palette',
    'coding': 'code',
    'translation': 'translate',
    'marketing': 'trending_up',
    'video': 'movie',
    'design': 'brush',
    'audio': 'music_note'
}

# Category colors
category_colors = {
    'writing': {'from': 'from-blue-500', 'to': 'to-cyan-600'},
    'creative': {'from': 'from-yellow-500', 'to': 'to-orange-600'},
    'office': {'from': 'from-green-500', 'to': 'to-emerald-600'},
    'image': {'from': 'from-purple-500', 'to': 'to-violet-600'},
    'learning': {'from': 'from-indigo-500', 'to': 'to-blue-600'},
    'chat': {'from': 'from-teal-500', 'to': 'to-cyan-600'},
    'toolbox': {'from': 'from-gray-500', 'to': 'to-slate-600'},
    'assistant': {'from': 'from-pink-500', 'to': 'to-rose-600'},
    'drawing': {'from': 'from-purple-500', 'to': 'to-pink-600'},
    'coding': {'from': 'from-emerald-500', 'to': 'to-green-600'},
    'translation': {'from': 'from-amber-500', 'to': 'to-yellow-600'},
    'marketing': {'from': 'from-red-500', 'to': 'to-pink-600'},
    'video': {'from': 'from-rose-500', 'to': 'to-red-600'},
    'design': {'from': 'from-fuchsia-500', 'to': 'to-purple-600'},
    'audio': {'from': 'from-sky-500', 'to': 'to-blue-600'}
}

all_tools = []
tools_by_category = {}

for sheet_name, cat_key in categories.items():
    ws = wb[sheet_name]
    tools_by_category[cat_key] = []
    
    for row in ws.iter_rows(min_row=2):
        tool = {
            'id': str(row[0].value or '').strip().replace(' ', '-').lower(),
            'name': str(row[0].value or '').strip(),
            'icon': str(row[1].value or '').strip(),
            'slogan': str(row[2].value or '').strip(),
            'rating': row[3].value,
            'level': str(row[4].value or '').strip(),
            'description': str(row[5].value or '').strip(),
            'tags': str(row[6].value or '').strip(),
            'monthlyTraffic': str(row[7].value or '').strip(),
            'intro': str(row[8].value or '').strip(),
            'targetAudience': str(row[9].value or '').strip(),
            'mainFeatures': str(row[10].value or '').strip(),
            'highlights': str(row[11].value or '').strip(),
            'status': str(row[12].value or '').strip(),
            'category': cat_key,
            'categoryName': category_names[cat_key]
        }
        
        all_tools.append(tool)
        tools_by_category[cat_key].append(tool)

# Generate category list for homepage
category_list = []
for cat_key in category_names.keys():
    if cat_key in tools_by_category:
        # Get top 7 for homepage
        top_tools = [{'name': t['name'], 'id': t['id'], 'icon': t['icon']} for t in tools_by_category[cat_key][:7]]
        category_list.append({
            'key': cat_key,
            'name': category_names[cat_key],
            'icon': category_icons[cat_key],
            'color': category_colors[cat_key],
            'count': len(tools_by_category[cat_key]),
            'topTools': top_tools
        })

# Generate API tools
wb_api = openpyxl.load_workbook('aixzd_aiapi.xlsx')
ws_api = wb_api.active

api_categories = {
    '聚合与路由平台': {'key': 'aggregation', 'icon': 'api', 'color': {'from': 'from-cyan-500', 'to': 'to-blue-600'}},
    '专项多模态与视觉': {'key': 'multimodal', 'icon': 'palette', 'color': {'from': 'from-pink-500', 'to': 'to-rose-600'}},
    '高性能极速推理': {'key': 'inference', 'icon': 'bolt', 'color': {'from': 'from-amber-500', 'to': 'to-orange-600'}},
    '旗舰模型原厂': {'key': 'flagship', 'icon': 'star', 'color': {'from': 'from-purple-500', 'to': 'to-indigo-600'}},
    '云巨头与企业级': {'key': 'enterprise', 'icon': 'cloud', 'color': {'from': 'from-emerald-500', 'to': 'to-teal-600'}},
    '企业级/开源网关': {'key': 'gateway', 'icon': 'shield', 'color': {'from': 'from-slate-500', 'to': 'to-gray-600'}}
}

api_tools = []
api_tools_by_category = {}

for row in ws_api.iter_rows(min_row=2):
    cat_name = str(row[2].value or '').strip()
    cat_info = api_categories.get(cat_name, {'key': 'other', 'icon': 'api', 'color': {'from': 'from-gray-500', 'to': 'to-slate-600'}})
    
    tool = {
        'id': str(row[1].value or '').strip().replace(' ', '-').replace('.', '').lower(),
        'rank': row[0].value,
        'name': str(row[1].value or '').strip(),
        'category': cat_name,
        'categoryKey': cat_info['key'],
        'description': str(row[3].value or '').strip(),
        'icon': '',  # Will be matched later
        'iconColor': cat_info['color'],
        'iconMaterial': cat_info['icon']
    }
    
    api_tools.append(tool)
    
    if cat_info['key'] not in api_tools_by_category:
        api_tools_by_category[cat_info['key']] = {
            'name': cat_name,
            'icon': cat_info['icon'],
            'color': cat_info['color'],
            'tools': []
        }
    api_tools_by_category[cat_info['key']]['tools'].append(tool)

# Try to match icons
product_icons = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        name = row[0].value
        icon_url = row[1].value
        if name and icon_url:
            key = str(name).strip().lower()
            product_icons[key] = str(icon_url).strip()

# Also read from ranking file
wb_rank = openpyxl.load_workbook('aixzd_ai排行表(1).xlsx')
for sheet_name in wb_rank.sheetnames:
    ws = wb_rank[sheet_name]
    for row in ws.iter_rows(min_row=2):
        name = row[1].value
        icon_url = row[8].value
        if name and icon_url:
            key = str(name).strip().lower()
            if key not in product_icons:
                product_icons[key] = str(icon_url).strip()

def find_icon(name):
    name_lower = name.lower().strip()
    
    def clean(s):
        return s.lower().replace(' ', '').replace('.', '').replace('-', '').replace('(', '').replace(')', '')
    
    clean_name = clean(name)
    
    # Exact match
    if name_lower in product_icons:
        return product_icons[name_lower]
    
    # Cleaned exact match
    for key, url in product_icons.items():
        if clean(key) == clean_name:
            return url
    
    # Try matching with common suffixes removed
    for suffix in [' api', ' ai', '.com', '.ai', '.io', ' (proxy)']:
        if clean_name.endswith(clean(suffix)):
            short_name = clean_name[:-len(clean(suffix))]
            for key, url in product_icons.items():
                if clean(key) == short_name:
                    return url
    
    # Word-by-word matching
    words = name_lower.split()
    if len(words) > 1:
        short_name = ' '.join(words[:-1])
        if short_name in product_icons:
            return product_icons[short_name]
    
    # Partial match
    for key, url in product_icons.items():
        key_clean = clean(key)
        if clean_name in key_clean and len(clean_name) >= 4:
            return url
        elif key_clean in clean_name and len(key_clean) >= 4:
            return url
    
    return None

# Match icons for API tools
for tool in api_tools:
    icon = find_icon(tool['name'])
    if icon:
        filename = icon.split('/')[-1]
        tool['icon'] = filename

# Save data
output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'src', 'data')
os.makedirs(output_dir, exist_ok=True)

# Save tools data
with open(os.path.join(output_dir, 'tools.js'), 'w', encoding='utf-8') as f:
    f.write('// Auto-generated tool data\n')
    f.write('export const tools = ')
    f.write(json.dumps(all_tools, ensure_ascii=False, indent=2))
    f.write(';\n')
    f.write('\nexport const toolsByCategory = ')
    f.write(json.dumps(tools_by_category, ensure_ascii=False, indent=2))
    f.write(';\n')
    f.write('\nexport const categoryList = ')
    f.write(json.dumps(category_list, ensure_ascii=False, indent=2))
    f.write(';\n')
    f.write('\nexport const categoryNames = ')
    f.write(json.dumps(category_names, ensure_ascii=False, indent=2))
    f.write(';\n')
    f.write('\nexport const categoryIcons = ')
    f.write(json.dumps(category_icons, ensure_ascii=False, indent=2))
    f.write(';\n')
    f.write('\nexport const categoryColors = ')
    f.write(json.dumps(category_colors, ensure_ascii=False, indent=2))
    f.write(';\n')

# Save API tools data
with open(os.path.join(output_dir, 'api.js'), 'w', encoding='utf-8') as f:
    f.write('// Auto-generated API tool data\n')
    f.write('export const apiTools = ')
    f.write(json.dumps(api_tools, ensure_ascii=False, indent=2))
    f.write(';\n')
    f.write('\nexport const apiToolsByCategory = ')
    f.write(json.dumps(api_tools_by_category, ensure_ascii=False, indent=2))
    f.write(';\n')

print(f'Total tools: {len(all_tools)}')
print(f'API tools: {len(api_tools)}')
print(f'Categories: {len(category_list)}')
print(f'\nGenerated files in {output_dir}')
